import openpyxl

from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from classySel import GeneralSearcher
from classyHTMLparser import ParseHTMLtables




def read_excel(filen):
    wb = openpyxl.load_workbook(filename = filen)
    row = 2
    addresscol = wb['sheet1'].cell(row=row, column=4).value
    
    a = GeneralSearcher()
    a.go_back_to_search_page()
    
    while addresscol != None:
        addy = addresscol
        ParseHTMLtables(a.search(addy))
        a.go_back_to_search_page()
        row+=1 
        addresscol = wb['sheet1'].cell(row=row, column=4).value
     



read_excel("propertyaddresses.xlsx")
